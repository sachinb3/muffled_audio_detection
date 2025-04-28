import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def highlight_results(input_file, output_file):
   """
   Highlight rows based on the comparison of actual muffling and guessed results.
   Args:
       input_file (str): Path to the input CSV or text file.
       output_file (str): Path to save the formatted Excel file.
   """
   df = pd.read_csv(input_file)

   wb = Workbook()
   ws = wb.active


   ws.append(list(df.columns))  # Write the header row
   for _, row in df.iterrows():
       ws.append(row.tolist())  # Write each row of data

   green_fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")  # Pastel green
   red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")    # Pastel red

   num_rows = df.shape[0]
   num_wrong = 0

   for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
       actual = row[3].value  # Column A (Actual)
       guessed = row[4].value  # Column B (Guessed)

       # Check the criteria for green or red highlighting
       if ("🟢" in actual and "CLEAR AUDIO" in guessed) or ("🔧" in actual and "MUFFLED AUDIO" in guessed):
           for cell in row:
               cell.fill = green_fill  # Highlight green
       else:
           num_wrong += 1
           for cell in row:
               cell.fill = red_fill  # Highlight red

   # Step 5: Save the formatted Excel file
   wb.save(output_file)
   print(f"✅ Results saved to {output_file}")
   accuracy = (num_rows - num_wrong) / num_rows
   print(f"❌ Accuracy = {accuracy}")


if __name__ == "__main__":
   # Input CSV file should have two columns: "Actual" (🟢 or 🔧) and "Guessed" (Clear Audio or Muffled Audio)
   input_file = "actual_muffling_log.csv"  # Replace with your actual input file path
   output_file = "results.xlsx"  # Replace with your desired output file path

   # Call the function to process and highlight the results
   highlight_results(input_file, output_file)

